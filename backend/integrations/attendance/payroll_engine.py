from __future__ import annotations

import calendar
from collections import defaultdict
from datetime import UTC, date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from database import models
from database.attendance_models import AttendanceRecord, AttendanceStatus, LeaveRequest, PayrollRecord
from database.attendance_schemas import CompanyPayrollSummaryOut, PayrollApprovalResult, PayrollRecordOut
from database.admin_crud import log_activity
from integrations.attendance.attendance_service import attendance_service

UZ_LABOR = {
    "standard_weekly_hours": 40,
    "standard_daily_hours": 8,
    "overtime_rate_first_2h": 1.5,
    "overtime_rate_after_2h": 2.0,
    "weekend_work_rate": 2.0,
    "inps_employee_rate": 0.04,
    "jshdssh_rate": 0.12,
    "min_wage_uzs": 980_000,
    "night_shift_start": 22,
    "night_shift_end": 6,
    "night_shift_premium": 0.20,
}


class PayrollEngine:
    @staticmethod
    def _utcnow() -> datetime:
        return datetime.now(UTC).replace(tzinfo=None)

    @staticmethod
    def _month_bounds(month: int, year: int) -> tuple[date, date]:
        start = date(year, month, 1)
        end = date(year, month, calendar.monthrange(year, month)[1])
        return start, end

    @staticmethod
    def _serialize_record(row: PayrollRecord, employee: models.Employee | None) -> PayrollRecordOut:
        return PayrollRecordOut(
            id=row.id,
            employee_id=row.employee_id,
            employee_name=employee.full_name if employee else f"Employee {row.employee_id}",
            department=employee.department if employee else None,
            position=employee.role if employee else None,
            company_id=row.company_id,
            period_month=row.period_month,
            period_year=row.period_year,
            working_days_in_month=row.working_days_in_month,
            days_worked=row.days_worked,
            days_absent=row.days_absent,
            days_on_leave=row.days_on_leave,
            total_hours_worked=float(row.total_hours_worked or 0),
            total_overtime_hours=float(row.total_overtime_hours or 0),
            total_late_minutes=int(row.total_late_minutes or 0),
            base_salary=float(row.base_salary or 0),
            prorated_salary=float(row.prorated_salary or 0),
            overtime_pay=float(row.overtime_pay or 0),
            late_penalty=float(row.late_penalty or 0),
            manual_penalty=float(row.manual_penalty or 0),
            bonus=float(row.bonus or 0),
            gross_salary=float(row.gross_salary or 0),
            inps_employee=float(row.inps_employee or 0),
            jshdssh=float(row.jshdssh or 0),
            total_deductions=float(row.total_deductions or 0),
            net_salary=float(row.net_salary or 0),
            status=row.status,
            approved_by=row.approved_by,
            approved_at=row.approved_at,
            paid_at=row.paid_at,
            payment_method=row.payment_method,
            is_manually_adjusted=bool(row.is_manually_adjusted),
            adjustment_note=row.adjustment_note,
            calculation_breakdown=row.calculation_breakdown or {},
            created_at=row.created_at,
            updated_at=row.updated_at,
        )

    def get_working_days_in_month(self, db: Session, company_id: int, month: int, year: int) -> int:
        start, end = self._month_bounds(month, year)
        current = start
        total = 0
        while current <= end:
            if attendance_service.is_working_day(db, current, company_id):
                total += 1
            current += timedelta(days=1)
        return total

    def _approved_leave_days(self, db: Session, employee: models.Employee, month: int, year: int) -> tuple[int, int]:
        start, end = self._month_bounds(month, year)
        rows = (
            db.query(LeaveRequest)
            .filter(
                LeaveRequest.employee_id == employee.id,
                LeaveRequest.company_id == int(employee.company_id or 0),
                LeaveRequest.status == "approved",
                LeaveRequest.date_from <= end,
                LeaveRequest.date_to >= start,
            )
            .all()
        )
        paid = 0
        unpaid = 0
        for row in rows:
            current = max(row.date_from, start)
            boundary = min(row.date_to, end)
            while current <= boundary:
                if attendance_service.is_working_day(db, current, int(employee.company_id or 0), list(employee.work_days or [1, 2, 3, 4, 5])):
                    if row.leave_type == "unpaid":
                        unpaid += 1
                    else:
                        paid += 1
                current += timedelta(days=1)
        return paid, unpaid

    def calculate_monthly_payroll(
        self,
        employee_id: int,
        month: int,
        year: int,
        db: Session,
        manual_bonus: float = 0,
        manual_penalty: float = 0,
        adjustment_note: str | None = None,
    ) -> PayrollRecord:
        employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
        if not employee or not employee.company_id:
            raise ValueError("Employee is missing company billing context.")
        start, end = self._month_bounds(month, year)
        attendance_rows = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.employee_id == employee.id,
                AttendanceRecord.work_date >= start,
                AttendanceRecord.work_date <= end,
            )
            .order_by(AttendanceRecord.work_date.asc())
            .all()
        )
        working_days = self.get_working_days_in_month(db, int(employee.company_id), month, year)
        paid_leave_days, unpaid_leave_days = self._approved_leave_days(db, employee, month, year)
        worked_rows = [row for row in attendance_rows if row.clock_in]
        days_worked = len(worked_rows)
        total_hours = round(sum(float(row.hours_worked or 0) for row in worked_rows), 2)
        total_overtime = round(sum(float(row.overtime_hours or 0) for row in worked_rows), 2)
        total_late_minutes = int(sum(int(row.late_minutes or 0) for row in worked_rows))
        raw_absent_days = working_days - days_worked - paid_leave_days
        days_absent = max(0, raw_absent_days)
        days_on_leave = paid_leave_days + unpaid_leave_days

        base_salary = float(employee.salary or 0)
        contract_type = (employee.contract_type or "monthly").strip().lower()
        hourly_rate = float(employee.hourly_rate or 0)
        hourly_equiv = hourly_rate if contract_type == "hourly" and hourly_rate > 0 else (base_salary / max(working_days, 1) / 8 if base_salary > 0 else 0)

        if contract_type == "hourly":
            prorated_salary = hourly_rate * total_hours
        elif contract_type == "daily":
            daily_rate = base_salary / max(working_days, 1)
            prorated_salary = daily_rate * days_worked
        else:
            prorated_salary = base_salary * ((days_worked + paid_leave_days) / max(working_days, 1))

        overtime_pay = 0.0
        for row in worked_rows:
            daily_ot = float(row.overtime_hours or 0)
            if daily_ot <= 0:
                continue
            rate_multiplier = UZ_LABOR["weekend_work_rate"] if not attendance_service.is_working_day(db, row.work_date, int(employee.company_id), list(employee.work_days or [1, 2, 3, 4, 5])) else None
            if rate_multiplier is not None:
                overtime_pay += daily_ot * hourly_equiv * rate_multiplier
                continue
            if daily_ot <= 2:
                overtime_pay += daily_ot * hourly_equiv * UZ_LABOR["overtime_rate_first_2h"]
            else:
                overtime_pay += 2 * hourly_equiv * UZ_LABOR["overtime_rate_first_2h"]
                overtime_pay += (daily_ot - 2) * hourly_equiv * UZ_LABOR["overtime_rate_after_2h"]
        overtime_pay = round(overtime_pay, 2)

        late_penalty = round((total_late_minutes / 60.0) * hourly_equiv, 2) if hourly_equiv > 0 else 0.0
        late_penalty = min(late_penalty, base_salary * 0.10 if base_salary > 0 else late_penalty)

        gross_salary = max(
            round(prorated_salary + overtime_pay + float(manual_bonus or 0) - late_penalty - float(manual_penalty or 0), 2),
            float(UZ_LABOR["min_wage_uzs"]),
        )
        inps_employee = round(gross_salary * UZ_LABOR["inps_employee_rate"], 2)
        taxable_income = gross_salary - inps_employee
        jshdssh = round(taxable_income * UZ_LABOR["jshdssh_rate"], 2)
        total_deductions = round(inps_employee + jshdssh, 2)
        net_salary = round(gross_salary - total_deductions, 2)

        breakdown = {
            "contract_type": contract_type,
            "base_salary": round(base_salary, 2),
            "working_days": working_days,
            "days_worked": days_worked,
            "paid_leave_days": paid_leave_days,
            "unpaid_leave_days": unpaid_leave_days,
            "total_hours_worked": total_hours,
            "total_overtime_hours": total_overtime,
            "hourly_equivalent": round(hourly_equiv, 2),
            "prorated_salary": round(prorated_salary, 2),
            "overtime_pay": overtime_pay,
            "late_penalty": late_penalty,
            "manual_bonus": round(float(manual_bonus or 0), 2),
            "manual_penalty": round(float(manual_penalty or 0), 2),
            "gross_salary": gross_salary,
            "inps_employee": inps_employee,
            "jshdssh": jshdssh,
            "net_salary": net_salary,
            "formula_lines": [
                f"Base salary: {base_salary:,.0f} UZS",
                f"Worked days: {days_worked}/{working_days} (+ paid leave {paid_leave_days})",
                f"Overtime pay: {overtime_pay:,.0f} UZS",
                f"INPS (4%): {inps_employee:,.0f} UZS",
                f"JShDSh (12%): {jshdssh:,.0f} UZS",
            ],
        }

        record = (
            db.query(PayrollRecord)
            .filter(
                PayrollRecord.employee_id == employee.id,
                PayrollRecord.period_month == month,
                PayrollRecord.period_year == year,
            )
            .first()
        )
        if not record:
            record = PayrollRecord(employee_id=employee.id, company_id=int(employee.company_id), period_month=month, period_year=year)
            db.add(record)
        record.working_days_in_month = working_days
        record.days_worked = days_worked
        record.days_absent = days_absent
        record.days_on_leave = days_on_leave
        record.total_hours_worked = total_hours
        record.total_overtime_hours = total_overtime
        record.total_late_minutes = total_late_minutes
        record.base_salary = round(base_salary, 2)
        record.prorated_salary = round(prorated_salary, 2)
        record.overtime_pay = overtime_pay
        record.late_penalty = late_penalty
        record.manual_penalty = round(float(manual_penalty or 0), 2)
        record.bonus = round(float(manual_bonus or 0), 2)
        record.gross_salary = gross_salary
        record.inps_employee = inps_employee
        record.jshdssh = jshdssh
        record.total_deductions = total_deductions
        record.net_salary = net_salary
        record.status = record.status if record.status in {"approved", "paid"} else "draft"
        record.is_manually_adjusted = bool(manual_bonus or manual_penalty or adjustment_note)
        record.adjustment_note = adjustment_note
        record.calculation_breakdown = breakdown
        db.commit()
        db.refresh(record)
        return record

    def calculate_company_payroll(self, company_id: int, month: int, year: int, db: Session, employee_ids: list[int] | None = None) -> CompanyPayrollSummaryOut:
        query = db.query(models.Employee).filter(models.Employee.company_id == company_id, models.Employee.status != models.EmployeeStatus.terminated)
        if employee_ids:
            query = query.filter(models.Employee.id.in_(employee_ids))
        employees = query.order_by(models.Employee.full_name).all()
        rows: list[PayrollRecordOut] = []
        warnings: list[str] = []
        total_gross = total_net = total_inps = total_tax = 0.0
        for employee in employees:
            record = self.calculate_monthly_payroll(employee.id, month, year, db)
            serialized = self._serialize_record(record, employee)
            rows.append(serialized)
            total_gross += serialized.gross_salary
            total_net += serialized.net_salary
            total_inps += serialized.inps_employee
            total_tax += serialized.jshdssh
            if serialized.days_absent > 0:
                warnings.append(f"{employee.full_name} has {serialized.days_absent} unexcused absence days.")
        return CompanyPayrollSummaryOut(
            records=rows,
            total_gross=round(total_gross, 2),
            total_net=round(total_net, 2),
            total_inps=round(total_inps, 2),
            total_jshdssh=round(total_tax, 2),
            total_deductions=round(total_inps + total_tax, 2),
            employee_count=len(rows),
            calculation_warnings=list(dict.fromkeys(warnings))[:10],
        )

    def list_payroll(self, db: Session, company_id: int, month: int, year: int) -> list[PayrollRecordOut]:
        rows = (
            db.query(PayrollRecord)
            .filter(PayrollRecord.company_id == company_id, PayrollRecord.period_month == month, PayrollRecord.period_year == year)
            .order_by(PayrollRecord.created_at.asc(), PayrollRecord.id.asc())
            .all()
        )
        employee_ids = [row.employee_id for row in rows]
        employees = {row.id: row for row in db.query(models.Employee).filter(models.Employee.id.in_(employee_ids)).all()} if employee_ids else {}
        return [self._serialize_record(row, employees.get(row.employee_id)) for row in rows]

    def adjust_payroll_record(self, db: Session, record: PayrollRecord, bonus: float | None, manual_penalty: float | None, adjustment_note: str | None) -> PayrollRecord:
        if record.status != "draft":
            raise ValueError("Only draft payroll records can be adjusted.")
        employee = db.query(models.Employee).filter(models.Employee.id == record.employee_id).first()
        if not employee:
            raise ValueError("Employee not found.")
        return self.calculate_monthly_payroll(
            employee_id=record.employee_id,
            month=record.period_month,
            year=record.period_year,
            db=db,
            manual_bonus=bonus if bonus is not None else float(record.bonus or 0),
            manual_penalty=manual_penalty if manual_penalty is not None else float(record.manual_penalty or 0),
            adjustment_note=adjustment_note if adjustment_note is not None else record.adjustment_note,
        )

    def approve_payroll(self, payroll_ids: list[int], approved_by: str, db: Session) -> PayrollApprovalResult:
        rows = db.query(PayrollRecord).filter(PayrollRecord.id.in_(payroll_ids)).all()
        approved_ids: list[int] = []
        company_ids: set[int] = set()
        for row in rows:
            if row.status == "paid":
                continue
            row.status = "approved"
            row.approved_by = approved_by
            row.approved_at = self._utcnow()
            approved_ids.append(row.id)
            company_ids.add(int(row.company_id))
        db.commit()
        for company_id in company_ids:
            log_activity(db, company_id, "payroll_approved", actor=approved_by, metadata=f"payroll_ids={approved_ids}")
        return PayrollApprovalResult(
            approved_count=len(approved_ids),
            payroll_ids=approved_ids,
            message=f"Approved {len(approved_ids)} payroll records.",
        )

    def export_payroll_excel(self, company_id: int, month: int, year: int, db: Session) -> bytes:
        rows = self.list_payroll(db, company_id, month, year)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        detail_sheet = workbook.create_sheet("Attendance Details")
        tax_sheet = workbook.create_sheet("Tax Summary")

        header_fill = PatternFill(fill_type="solid", fgColor="7C6AFF")
        header_font = Font(color="FFFFFF", bold=True)

        summary_headers = ["Employee Name", "Department", "Position", "Days Worked", "Hours", "OT Hours", "Base", "Gross", "INPS", "JShDSh", "Net", "Status"]
        summary_sheet.append(summary_headers)
        for cell in summary_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            summary_sheet.append([
                row.employee_name,
                row.department or "",
                row.position or "",
                row.days_worked,
                row.total_hours_worked,
                row.total_overtime_hours,
                row.base_salary,
                row.gross_salary,
                row.inps_employee,
                row.jshdssh,
                row.net_salary,
                row.status,
            ])

        detail_headers = ["Employee", "Days Worked", "Absent", "Leave", "Late Minutes", "Overtime Hours", "Breakdown"]
        detail_sheet.append(detail_headers)
        for cell in detail_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            detail_sheet.append([
                row.employee_name,
                row.days_worked,
                row.days_absent,
                row.days_on_leave,
                row.total_late_minutes,
                row.total_overtime_hours,
                " | ".join(row.calculation_breakdown.get("formula_lines", [])),
            ])

        tax_headers = ["Employee", "Gross Salary", "INPS", "JShDSh", "Total Deductions", "Net Salary"]
        tax_sheet.append(tax_headers)
        for cell in tax_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            tax_sheet.append([
                row.employee_name,
                row.gross_salary,
                row.inps_employee,
                row.jshdssh,
                row.total_deductions,
                row.net_salary,
            ])

        stream = __import__("io").BytesIO()
        workbook.save(stream)
        return stream.getvalue()


payroll_engine = PayrollEngine()
